"""
聯博投信 (AllianceBernstein) ETF 爬蟲模組

聯博官網 ETF「申購買回清單」頁面的資料由 webapi.alliancebernstein.com 提供，
且官方提供「基金持股權重」xlsx 下載 (法規必揭、最穩定的資料點)，故沿用與
[[morgan_scraper]] 相同的「下載檔案 -> openpyxl 解析」策略。

xlsx 下載端點 (從官網 clientlib JS 反推)：
  https://webapi.alliancebernstein.com/v2/funds/tw/zh-tw/investor/<shareClassId>/xls?type=holdings&prec=2
回傳 content-disposition: attachment; filename=YYYY-MM-DD-Portfolio-Holding.xlsx，
檔案資料日期即取自該檔名。

注意 (與 src/allianz_scraper.py 的「安聯 Allianz」不同公司)：本檔為「聯博 AllianceBernstein」。
"""
import io
import re
import random
import time
from typing import List, Dict, Any, Optional

import openpyxl
import requests
import urllib3
from loguru import logger
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from .config import REQUEST_DELAY_MIN, REQUEST_DELAY_MAX, MAX_RETRIES


# 聯博投信 ETF 與其 shareClassId 對照表
AB_ETF_CODES = {
    '00404A': 'TW00000404A5',  # 聯博台灣動能收益50主動式ETF
}

XLS_URL_TEMPLATE = (
    'https://webapi.alliancebernstein.com/v2/funds/tw/zh-tw/investor/'
    '{share_class_id}/xls?type=holdings&prec=2'
)
PCF_PAGE_TEMPLATE = 'https://www.abfunds.com.tw/zh-tw/etfs/pcf.{share_class_id}.html'

# 非台股 ISIN (KY/海外註冊但於台灣掛牌) 無法由 ISIN 還原台股代號，需人工對照。
# 未列入者會在解析時 WARN，並暫以原始 ISIN 當代號存入 (避免遺漏導致權重不齊)。
AB_ISIN_OVERRIDE = {
    'KYG9721M1033': '2637',  # 慧洋-KY
}


class ABFundsScraper:
    """聯博投信 ETF 爬蟲 (透過 holdings xlsx)"""

    def __init__(self):
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        session = requests.Session()
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        retry = Retry(
            total=MAX_RETRIES,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        return session

    def _headers(self, share_class_id: str) -> Dict[str, str]:
        return {
            'User-Agent': (
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            ),
            'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
            'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8',
            'Referer': PCF_PAGE_TEMPLATE.format(share_class_id=share_class_id),
        }

    def _random_delay(self):
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    def get_all_mappings(self) -> Dict[str, str]:
        return AB_ETF_CODES.copy()

    @staticmethod
    def _isin_to_stock_code(isin: str) -> Optional[str]:
        """台股 ISIN (TW000 + 4碼代號 + 檢查碼) 取代號；非台股查 override。"""
        isin = (isin or '').strip()
        if isin.startswith('TW') and len(isin) >= 9 and isin[5:9].isdigit():
            return isin[5:9]
        return AB_ISIN_OVERRIDE.get(isin)

    @staticmethod
    def _to_int(raw: Any) -> int:
        try:
            return int(float(str(raw).replace(',', ''))) if raw is not None else 0
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _to_float(raw: Any) -> float:
        try:
            return float(str(raw).replace(',', '').replace('%', '')) if raw is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _download_xls(self, etf_code: str, share_class_id: str) -> Optional[requests.Response]:
        url = XLS_URL_TEMPLATE.format(share_class_id=share_class_id)
        logger.info(f"Downloading AB holdings xlsx: {url}")
        self._random_delay()
        try:
            r = self.session.get(url, headers=self._headers(share_class_id), timeout=30, verify=False)
            r.raise_for_status()
            if not r.content or r.content[:2] != b'PK':
                logger.error(f"{etf_code}: response is not a valid xlsx (magic={r.content[:4]!r})")
                return None
            return r
        except requests.exceptions.RequestException as e:
            logger.error(f"AB holdings xlsx download failed for {etf_code}: {e}")
            return None

    @staticmethod
    def _date_from_response(resp: requests.Response, fallback: str) -> str:
        cd = resp.headers.get('content-disposition', '')
        m = re.search(r'(\d{4}-\d{2}-\d{2})', cd)
        return m.group(1) if m else fallback

    def get_etf_holdings(self, etf_code: str, date: str) -> List[Dict[str, Any]]:
        """從 holdings xlsx 解析持股清單。

        傳入 date 僅為期望目標日；實際 date 取自下載檔名 (content-disposition)。
        """
        share_class_id = AB_ETF_CODES.get(etf_code)
        if not share_class_id:
            logger.error(f"AB: ETF {etf_code} not in mapping")
            return []

        resp = self._download_xls(etf_code, share_class_id)
        if not resp:
            return []

        actual_date = self._date_from_response(resp, date)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
        except Exception as e:
            logger.error(f"AB: cannot parse xlsx for {etf_code}: {e}")
            return []

        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))

        # 定位股票持股表：表頭為「代碼/名稱/股數/市值/權重」(期貨/期權表用「口數」故排除)。
        header_idx = None
        for i, r in enumerate(rows):
            cells = [str(c).strip() if c is not None else '' for c in r[:5]]
            if '代碼' in cells and '名稱' in cells and '股數' in cells and '權重' in cells:
                header_idx = i
                break
        if header_idx is None:
            logger.warning(f"AB: stock holdings table not found for {etf_code} (xlsx structure changed?)")
            return []

        holdings = []
        for r in rows[header_idx + 1:]:
            if not r or r[0] is None or str(r[0]).strip() == '':
                break  # 空列代表股票段結束 (後面是期貨/期權/債券段)
            isin = str(r[0]).strip()
            name = (str(r[1]).strip() if r[1] is not None else '')
            shares = self._to_int(r[2])
            market_value = self._to_float(r[3])
            weight = self._to_float(r[4])

            stock_code = self._isin_to_stock_code(isin)
            if not stock_code:
                logger.warning(f"AB: {etf_code} unmapped non-TW ISIN '{isin}' ({name}); storing raw ISIN as code")
                stock_code = isin

            holdings.append({
                'etf_code': etf_code,
                'stock_code': stock_code,
                'stock_name': name,
                'shares': shares,
                'market_value': market_value,
                'weight': weight,
                'date': actual_date,
            })

        logger.info(f"AB: parsed {len(holdings)} holdings for {etf_code} on {actual_date}")
        return holdings
