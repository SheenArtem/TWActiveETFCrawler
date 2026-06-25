"""
摩根投信 ETF 爬蟲模組
直接下載 J.P. Morgan AM 在 AEM DAM 上發佈的 PCF (申購買回清單) Excel，
從中解析出持股股數、市值、估值日期，依市值反推權重。
"""
import io
import random
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import openpyxl
import requests
from loguru import logger
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from .config import REQUEST_DELAY_MIN, REQUEST_DELAY_MAX, MAX_RETRIES


# 摩根投信 ETF 與其 ISIN (CUSIP) 對照表
# PCF Excel URL 規則：
#   https://am.jpmorgan.com/content/dam/jpm-am-aem/asiapacific/tw/zh/
#   regulatory/etf-supplement/jpm_apac_tw_etf_pcf_updates_<ETF>_<ISIN>.xlsx
MORGAN_ETF_CODES = {
    '00401A': 'TW00000401A1',  # 主動摩根台灣鑫收
}

PCF_URL_TEMPLATE = (
    'https://am.jpmorgan.com/content/dam/jpm-am-aem/asiapacific/tw/zh/'
    'regulatory/etf-supplement/jpm_apac_tw_etf_pcf_updates_{etf_code}_{isin}.xlsx'
)
PRODUCT_PAGE_TEMPLATE = (
    'https://am.jpmorgan.com/tw/zh/asset-management/twetf/products/'
    'jpmorgan-taiwan-taiwan-equity-high-income-active-etf-{isin_lower}'
)


class MorganScraper:
    """摩根投信 ETF 爬蟲 (透過 PCF xlsx)"""

    def __init__(self):
        self.session = self._create_session()
        self.download_dir = Path('downloads/morgan')
        self.download_dir.mkdir(parents=True, exist_ok=True)

    def _create_session(self) -> requests.Session:
        session = requests.Session()
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        retry = Retry(
            total=MAX_RETRIES,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET"],
        )
        adapter = HTTPAdapter(max_retries=retry)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        return session

    def _headers(self, isin: str) -> Dict[str, str]:
        return {
            'User-Agent': (
                'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            ),
            'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*',
            'Accept-Language': 'zh-TW,zh;q=0.9,en;q=0.8',
            'Referer': PRODUCT_PAGE_TEMPLATE.format(isin_lower=isin.lower()),
        }

    def _random_delay(self):
        delay = random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX)
        time.sleep(delay)

    def get_all_mappings(self) -> Dict[str, str]:
        return MORGAN_ETF_CODES.copy()

    def _download_pcf(self, etf_code: str, isin: str) -> Optional[bytes]:
        url = PCF_URL_TEMPLATE.format(etf_code=etf_code, isin=isin)
        logger.info(f"Downloading Morgan PCF: {url}")
        self._random_delay()
        try:
            r = self.session.get(url, headers=self._headers(isin), timeout=30)
            r.raise_for_status()
            if not r.content or r.content[:2] != b'PK':
                logger.error(f"{etf_code}: response is not a valid xlsx (magic={r.content[:4]!r})")
                return None
            # 同步存一份做備查
            save_path = self.download_dir / f'{etf_code}_pcf.xlsx'
            save_path.write_bytes(r.content)
            logger.debug(f"Saved PCF to {save_path}")
            return r.content
        except requests.exceptions.RequestException as e:
            logger.error(f"Morgan PCF download failed for {etf_code}: {e}")
            return None

    @staticmethod
    def _parse_valuation_date(raw: Any) -> str:
        """將 PCF 的 Valuation Date (YYYYMMDD) 轉成 YYYY-MM-DD。"""
        if raw is None:
            return ''
        s = str(raw).strip()
        if len(s) == 8 and s.isdigit():
            return f'{s[0:4]}-{s[4:6]}-{s[6:8]}'
        return s

    def get_etf_holdings(self, etf_code: str, date: str) -> List[Dict[str, Any]]:
        """
        從 PCF xlsx 解析持股清單。

        傳入 date 僅作為呼叫端期望的目標日期；實際 date 以 xlsx 中的
        Valuation Date 為準（會覆寫 holding['date']）。
        """
        isin = MORGAN_ETF_CODES.get(etf_code)
        if not isin:
            logger.error(f"Morgan: ETF {etf_code} not in mapping")
            return []

        content = self._download_pcf(etf_code, isin)
        if not content:
            return []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            logger.error(f"Morgan: cannot parse xlsx for {etf_code}: {e}")
            return []

        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            logger.warning(f"Morgan: xlsx for {etf_code} has too few rows ({len(rows)})")
            return []

        # 第 1 列：fund-level 欄位名
        # 第 2 列：fund-level 資料 (Record Type='H')
        # 第 3 列：constituent 欄位名
        # 第 4 列起：D 列為個股
        fund_header = rows[0]
        fund_data = rows[1]
        fund_row = dict(zip(fund_header, fund_data))
        valuation_date = self._parse_valuation_date(fund_row.get('Valuation Date'))
        estimated_total_mv = fund_row.get('Estimated Total Market Value') or 0
        if not valuation_date:
            valuation_date = date  # fallback
        # JPMorgan PCF (申購買回清單) 的 Valuation Date 是「次一交易日」(申購買回基準日)，
        # 並非持股 as-of 日。若解析出的估值日晚於請求日(今天的交易日)，代表這是隔日的
        # 前瞻性籃子，應以請求日作為持股日期，避免把未來日期寫進 DB 造成報表/網頁日期超前。
        if valuation_date > date:
            logger.info(
                f"Morgan {etf_code}: PCF valuation date {valuation_date} is later than "
                f"request date {date} (forward-looking basket); storing as {date}"
            )
            valuation_date = date

        # 找 constituent 欄位 index
        cons_header = rows[2]
        col = {name: idx for idx, name in enumerate(cons_header) if name}
        required = ('Constituent Ticker', 'Constituent Description',
                    'Shares or PAR Amount', 'Market Value Base', 'Record Type')
        for k in required:
            if k not in col:
                logger.error(f"Morgan: missing column '{k}' in PCF xlsx for {etf_code}")
                return []

        holdings = []
        for raw in rows[3:]:
            if not raw or raw[col['Record Type']] != 'D':
                continue
            ticker = raw[col['Constituent Ticker']]
            if not ticker:
                continue
            stock_code = str(ticker).strip()
            stock_name = (raw[col['Constituent Description']] or '').strip()
            shares_raw = raw[col['Shares or PAR Amount']]
            mv_raw = raw[col['Market Value Base']]
            try:
                shares = int(float(shares_raw)) if shares_raw is not None else 0
            except (TypeError, ValueError):
                shares = 0
            try:
                market_value = float(mv_raw) if mv_raw is not None else 0.0
            except (TypeError, ValueError):
                market_value = 0.0
            weight = 0.0
            if estimated_total_mv:
                try:
                    weight = round(market_value / float(estimated_total_mv) * 100, 4)
                except (TypeError, ValueError, ZeroDivisionError):
                    weight = 0.0
            holdings.append({
                'etf_code': etf_code,
                'stock_code': stock_code,
                'stock_name': stock_name,
                'shares': shares,
                'market_value': market_value,
                'weight': weight,
                'date': valuation_date,
            })

        logger.info(f"Morgan: parsed {len(holdings)} holdings for {etf_code} on {valuation_date}")
        return holdings
